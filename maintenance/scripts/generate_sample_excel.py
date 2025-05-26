import openpyxl
from openpyxl.styles import Font

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active

# Set column headers
headers = ['Code', 'Item', 'Machine', 'Inventory']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

# Add sample data
data = [
    ['SPB002', 'LOAD CELL FOR MIXER 1.1 TON', 'Dosing', 3],
    ['SPB003', 'SHAFT', 'Dosing', 3],
    ['SPB004', 'Bearing NTN NJ222', 'Dosing', 2],
    ['SPB005', 'Mixer Push-button', 'Dosing', 9],
    ['SPB006', 'Main shaft bearing', 'Dosing', 1]
]

# Add data rows
for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Save the workbook
wb.save('sample_spare_parts.xlsx')
print('Sample Excel file created successfully!')
