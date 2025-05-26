from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import admin
from django.contrib.auth import views as auth_views
from django.http import HttpResponse
from django.contrib import messages
from django.db import models
from django.db.models import Count, Q, Sum, Max
from django.db.models.functions import ExtractMonth, ExtractYear
from django.utils import timezone
from django.urls import path, include, reverse
from django.template.loader import get_template
from xhtml2pdf import pisa
import xlsxwriter
from io import BytesIO
import json
import calendar
from django.contrib.auth.models import User

from .models import Machine, UserProfile, UserRole, Fault, FaultPhoto, SparePart, SparePartTransaction, PurchaseRequest, PurchaseItem
from .forms import SparePartImportForm, SparePartTransactionForm, SparePartForm, TransactionFilterForm, SparePartsFilterForm, PurchaseRequestForm, PurchaseItemForm, PurchaseFilterForm, PurchaseItemEditForm

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import xlsxwriter
import openpyxl
from django.conf import settings
from django.conf.urls.static import static

@login_required
def dashboard(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        # Create default admin role if it doesn't exist
        admin_role, _ = UserRole.objects.get_or_create(
            name='admin',
            defaults={
                'can_add_users': True,
                'can_delete_users': True,
                'can_view_all': True,
                'can_export': True
            }
        )
        # Create user profile
        user_profile = UserProfile.objects.create(user=user, role=admin_role)

    # Get all machines
    machines = Machine.objects.all()
    total_machines = machines.count()
    
    # Get recent faults
    recent_faults = Fault.objects.order_by('-reported_at')[:5]
    total_faults = Fault.objects.count()
    active_faults = Fault.objects.filter(machine__status='maintenance').count()
    
    # Get top spare parts by consumption
    top_spare_parts = SparePart.objects.annotate(
        total_consumption=Count('transactions', filter=Q(transactions__transaction_type='remove'))
    ).order_by('-total_consumption')[:5]
    
    # Prepare data for top spare parts chart
    top_spare_parts_labels = [sp.code for sp in top_spare_parts]
    top_spare_parts_data = [sp.total_consumption for sp in top_spare_parts]
    
    # Get pending purchase requests
    pending_requests = PurchaseRequest.objects.filter(
        request_type='pending'
    ).order_by('-request_date')[:5]
    
    # Get monthly expenses
    monthly_expenses = PurchaseRequest.objects.filter(
        request_type='approved'
    ).annotate(
        month=ExtractMonth('request_date'),
        year=ExtractYear('request_date')
    ).values('month', 'year').annotate(
        total=Sum('total_price')
    ).order_by('year', 'month')
    
    # Prepare data for monthly expenses chart
    monthly_expenses_labels = []
    monthly_expenses_data = []
    for expense in monthly_expenses:
        month_name = calendar.month_name[expense['month']]
        monthly_expenses_labels.append(f"{month_name} {expense['year']}")
        monthly_expenses_data.append(float(expense['total']))
    
    # Get most frequent breakdowns
    frequent_breakdowns = Machine.objects.annotate(
        breakdown_count=Count('fault'),
        last_breakdown=Max('fault__reported_at', output_field=models.DateTimeField())
    ).filter(
        breakdown_count__gt=0
    ).order_by('-breakdown_count')[:5]
    
    context = {
        'user_profile': user_profile,
        'machines': machines,
        'total_machines': total_machines,
        'active_faults': active_faults,
        'total_faults': total_faults,
        'recent_faults': recent_faults,
        'pending_requests': pending_requests,
        'frequent_breakdowns': frequent_breakdowns,
        'top_spare_parts_labels': json.dumps(top_spare_parts_labels),
        'top_spare_parts_data': json.dumps(top_spare_parts_data),
        'monthly_expenses_labels': json.dumps(monthly_expenses_labels),
        'monthly_expenses_data': json.dumps(monthly_expenses_data)
    }
    return render(request, 'maintenance/dashboard.html', context)

@login_required
def machine_list(request):
    machines = Machine.objects.all()
    return render(request, 'maintenance/machine_list.html', {'machines': machines})

@login_required
def machine_detail(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    return render(request, 'maintenance/machine_detail.html', {'machine': machine})

@login_required
def machine_delete(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        machine.delete()
        messages.success(request, 'Machine deleted successfully.')
        return redirect('maintenance:machine_list')
    return render(request, 'maintenance/machine_confirm_delete.html', {'machine': machine})

@login_required
def spare_parts_list(request):
    filter_form = SparePartsFilterForm(request.GET)
    spare_parts = SparePart.objects.all()
    
    if filter_form.is_valid():
        start_date = filter_form.cleaned_data['start_date']
        end_date = filter_form.cleaned_data['end_date']
        spare_part_id = filter_form.cleaned_data['spare_part']
        
        if start_date:
            spare_parts = spare_parts.filter(transactions__transaction_date__gte=start_date)
        if end_date:
            spare_parts = spare_parts.filter(transactions__transaction_date__lte=end_date)
        if spare_part_id:
            spare_parts = spare_parts.filter(id=spare_part_id)

    context = {
        'spare_parts': spare_parts,
        'filter_form': filter_form,
        'machines': Machine.objects.all(),
        'machines_list': Machine.objects.all()  # For transaction form
    }
    return render(request, 'maintenance/spare_parts.html', context)

@login_required
def add_transaction(request, spare_part_id):
    spare_part = get_object_or_404(SparePart, id=spare_part_id)
    
    if request.method == 'POST':
        form = SparePartTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.spare_part = spare_part
            transaction.created_by = request.user
            
            # Update inventory based on transaction type
            if transaction.transaction_type == 'add':
                spare_part.current_inventory += transaction.quantity
            else:  # remove
                spare_part.current_inventory -= transaction.quantity
            
            spare_part.save()
            transaction.save()
            
            messages.success(request, 'Transaction added successfully')
            return redirect('maintenance:spare_parts_list')
    
    form = SparePartTransactionForm()
    context = {
        'form': form,
        'spare_part': spare_part,
        'machines': Machine.objects.all()
    }
    return render(request, 'maintenance/add_transaction.html', context)

@login_required
def spare_part_detail(request, pk):
    spare_part = get_object_or_404(SparePart, pk=pk)
    transactions = spare_part.transactions.all().order_by('-transaction_date')
    
    if request.method == 'POST':
        form = SparePartTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.spare_part = spare_part
            transaction.created_by = request.user
            
            # Update inventory based on transaction type
            if transaction.transaction_type == 'add':
                spare_part.current_inventory += transaction.quantity
            else:  # remove
                spare_part.current_inventory -= transaction.quantity
            
            spare_part.save()
            transaction.save()
            
            messages.success(request, 'Transaction added successfully')
            return redirect('maintenance:spare_part_detail', pk=pk)
    else:
        form = SparePartTransactionForm()
    
    context = {
        'spare_part': spare_part,
        'transactions': transactions,
        'form': form
    }
    return render(request, 'maintenance/spare_part_detail.html', context)

@login_required
def add_transaction(request, spare_part_id):
    spare_part = get_object_or_404(SparePart, id=spare_part_id)
    
    if request.method == 'POST':
        form = SparePartTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.spare_part = spare_part
            transaction.save()
            
            # Update inventory based on transaction type
            if transaction.transaction_type == 'add':
                spare_part.current_inventory += transaction.quantity
            else:  # remove
                spare_part.current_inventory -= transaction.quantity
            spare_part.save()
            
            messages.success(request, 'Transaction added successfully.')
            return redirect('maintenance:spare_part_detail', pk=spare_part_id)
    else:
        form = SparePartTransactionForm()
    
    context = {
        'form': form,
        'spare_part': spare_part
    }
    return render(request, 'maintenance/transaction_form.html', context)

@login_required
def purchase_list(request):
    filter_form = PurchaseFilterForm(request.GET)
    purchases = PurchaseRequest.objects.all().order_by('-request_date')
    
    if filter_form.is_valid():
        start_date = filter_form.cleaned_data['start_date']
        end_date = filter_form.cleaned_data['end_date']
        
        if start_date:
            purchases = purchases.filter(request_date__gte=start_date)
        if end_date:
            purchases = purchases.filter(request_date__lte=end_date)
    
    context = {
        'purchases': purchases,
        'filter_form': filter_form
    }
    return render(request, 'maintenance/purchase_list.html', context)

@login_required
def purchase_add(request):
    if request.method == 'POST':
        form = PurchaseRequestForm(request.POST)
        if form.is_valid():
            purchase = form.save(commit=False)
            purchase.created_by = request.user
            purchase.save()
            messages.success(request, 'Purchase request created successfully.')
            return redirect('maintenance:purchase_detail', pk=purchase.pk)
    else:
        form = PurchaseRequestForm()
    
    context = {
        'form': form,
        'title': 'Add Purchase Request'
    }
    return render(request, 'maintenance/purchase_form.html', context)

@login_required
def purchase_edit(request, pk):
    purchase = get_object_or_404(PurchaseRequest, pk=pk)
    
    if request.method == 'POST':
        form = PurchaseRequestForm(request.POST, instance=purchase)
        if form.is_valid():
            form.save()
            messages.success(request, 'Purchase request updated successfully.')
            return redirect('maintenance:purchase_detail', pk=purchase.pk)
    else:
        form = PurchaseRequestForm(instance=purchase)
    
    context = {
        'form': form,
        'title': 'Edit Purchase Request'
    }
    return render(request, 'maintenance/purchase_form.html', context)

@login_required
def purchase_detail(request, pk):
    purchase = get_object_or_404(PurchaseRequest, pk=pk)
    items = purchase.items.all()
    
    if request.method == 'POST':
        form = PurchaseItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.purchase_request = purchase
            item.save()
            messages.success(request, 'Item added successfully.')
            return redirect('maintenance:purchase_detail', pk=pk)
    else:
        form = PurchaseItemForm()
    
    # Calculate total price
    total_price = purchase.calculate_total_price()
    
    context = {
        'purchase': purchase,
        'items': items,
        'form': form,
        'total_price': total_price
    }
    return render(request, 'maintenance/purchase_detail.html', context)

@login_required
def purchase_item_delete(request, purchase_id, item_id):
    purchase = get_object_or_404(PurchaseRequest, pk=purchase_id)
    item = get_object_or_404(PurchaseItem, pk=item_id)
    
    if request.method == 'POST':
        item.delete()
        messages.success(request, 'Item deleted successfully.')
        return redirect('maintenance:purchase_detail', pk=purchase_id)
    
    return render(request, 'maintenance/purchase_item_confirm_delete.html', {
        'item': item,
        'purchase': purchase
    })

@login_required
def purchase_item_edit(request, purchase_id, item_id):
    purchase = get_object_or_404(PurchaseRequest, pk=purchase_id)
    item = get_object_or_404(PurchaseItem, pk=item_id)
    
    if request.method == 'POST':
        form = PurchaseItemEditForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Item updated successfully.')
            return redirect('maintenance:purchase_detail', pk=purchase_id)
    else:
        form = PurchaseItemEditForm(instance=item)
    
    # Calculate total price for display
    total_price = purchase.calculate_total_price()
    
    context = {
        'form': form,
        'purchase': purchase,
        'item': item,
        'total_price': total_price
    }
    return render(request, 'maintenance/purchase_item_edit.html', context)

@login_required
def purchase_delete(request, pk):
    purchase = get_object_or_404(PurchaseRequest, pk=pk)
    if request.method == 'POST':
        purchase.delete()
        messages.success(request, 'Purchase request deleted successfully.')
        return redirect('maintenance:purchase_list')
    return render(request, 'maintenance/purchase_confirm_delete.html', {'purchase': purchase})

@login_required
def transaction_list(request):
    filter_form = TransactionFilterForm(request.GET)
    transactions = SparePartTransaction.objects.all().order_by('-transaction_date')
    
    if filter_form.is_valid():
        start_date = filter_form.cleaned_data['start_date']
        end_date = filter_form.cleaned_data['end_date']
        spare_part = filter_form.cleaned_data['spare_part']
        
        if start_date:
            transactions = transactions.filter(transaction_date__gte=start_date)
        if end_date:
            transactions = transactions.filter(transaction_date__lte=end_date)
        if spare_part:
            transactions = transactions.filter(spare_part=spare_part)
    
    context = {
        'transactions': transactions,
        'filter_form': filter_form
    }
    return render(request, 'maintenance/transaction_list.html', context)

@login_required
def export_purchases_excel(request):
    purchases = PurchaseRequest.objects.all().order_by('-request_date')
    
    # Create a new workbook and add a worksheet
    workbook = xlsxwriter.Workbook('purchases.xlsx')
    worksheet = workbook.add_worksheet()
    
    # Add headers
    headers = [
        'Request Number', 'Request Type', 'Request Date', 'Submission Date',
        'Item Count', 'Total Price', 'Additional Info'
    ]
    
    row = 0
    col = 0
    
    # Write headers
    for header in headers:
        worksheet.write(row, col, header)
        col += 1
    
    # Write data
    row = 1
    for purchase in purchases:
        worksheet.write(row, 0, purchase.request_number)
        worksheet.write(row, 1, purchase.get_request_type_display())
        worksheet.write(row, 2, purchase.request_date.strftime('%Y-%m-%d'))
        worksheet.write(row, 3, purchase.submission_date.strftime('%Y-%m-%d'))
        worksheet.write(row, 4, purchase.get_item_count())
        worksheet.write(row, 5, float(purchase.total_price))
        worksheet.write(row, 6, purchase.additional_info)
        row += 1
    
    workbook.close()
    
    # Serve the file
    with open('purchases.xlsx', 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=purchases.xlsx'
    return response

@login_required
def export_purchases_pdf(request):
    purchases = PurchaseRequest.objects.all().order_by('-request_date')
    
    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="purchases.pdf"'
    
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    # Add title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "Purchase Requests Report")
    
    # Add date
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 70, f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Add table header
    headers = [
        'Request Number', 'Request Type', 'Request Date', 'Submission Date',
        'Item Count', 'Total Price', 'Additional Info'
    ]
    
    # Draw table
    y = height - 100
    for header in headers:
        p.drawString(50, y, header)
        y -= 20
    
    # Draw data
    for purchase in purchases:
        y -= 20
        if y < 50:
            p.showPage()
            y = height - 50
        
        p.drawString(50, y, purchase.request_number)
        p.drawString(150, y, purchase.get_request_type_display())
        p.drawString(250, y, purchase.request_date.strftime('%Y-%m-%d'))
        p.drawString(350, y, purchase.submission_date.strftime('%Y-%m-%d'))
        p.drawString(450, y, str(purchase.get_item_count()))
        p.drawString(500, y, f"{purchase.total_price:.2f}")
        p.drawString(550, y, purchase.additional_info or '')
    
    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()
    
    return response

@login_required
def export_transactions_excel(request, transactions):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="transactions_{}.xlsx"'.format(timezone.now().strftime('%Y%m%d'))
    
    workbook = xlsxwriter.Workbook(response, {'in_memory': True})
    worksheet = workbook.add_worksheet()
    
    # Add headers
    headers = ['Date', 'Spare Part Code', 'Spare Part Item', 'Transaction Type', 'Quantity', 'Machine', 'Notes']
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header)
    
    # Add data
    row = 1
    for transaction in transactions:
        worksheet.write(row, 0, transaction.transaction_date.strftime('%Y-%m-%d %H:%M'))
        worksheet.write(row, 1, transaction.spare_part.code)
        worksheet.write(row, 2, transaction.spare_part.item)
        worksheet.write(row, 3, transaction.get_transaction_type_display())
        worksheet.write(row, 4, transaction.quantity)
        worksheet.write(row, 5, transaction.machine.name if transaction.machine else '-')
        worksheet.write(row, 6, transaction.notes)
        row += 1
    
    workbook.close()
    return response

@login_required
def export_transactions_pdf(request, transactions):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="transactions_{}.pdf"'.format(timezone.now().strftime('%Y%m%d'))
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title = Paragraph('Spare Part Transactions', styles['Title'])
    elements.append(title)
    
    # Add date range if filtered
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date or end_date:
        date_range = Paragraph('Date Range: {} to {}'.format(
            start_date or 'Start Date',
            end_date or 'End Date'
        ), styles['Normal'])
        elements.append(date_range)
    
    # Add table
    headers = ['Date', 'Spare Part Code', 'Spare Part Item', 'Transaction Type', 'Quantity', 'Machine', 'Notes']
    data = [headers]
    
    for transaction in transactions:
        row = [
            transaction.transaction_date.strftime('%Y-%m-%d %H:%M'),
            transaction.spare_part.code,
            transaction.spare_part.item,
            transaction.get_transaction_type_display(),
            str(transaction.quantity),
            transaction.machine.name if transaction.machine else '-',
            transaction.notes
        ]
        data.append(row)
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 14),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 12),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response

@login_required
def export_spare_parts_excel(request):
    spare_parts = SparePart.objects.all()
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="spare_parts_inventory_{}.xlsx"'.format(timezone.now().strftime('%Y%m%d'))
    
    workbook = xlsxwriter.Workbook(response, {'in_memory': True})
    worksheet = workbook.add_worksheet()
    
    # Add headers
    headers = ['Code', 'Item', 'Current Inventory']
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header)
    
    # Add data
    row = 1
    for spare_part in spare_parts:
        worksheet.write(row, 0, spare_part.code)
        worksheet.write(row, 1, spare_part.item)
        worksheet.write(row, 2, spare_part.inventory)
        row += 1
    
    workbook.close()
    return response

@login_required
def export_spare_parts_pdf(request):
    spare_parts = SparePart.objects.all()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="spare_parts_inventory_{}.pdf"'.format(timezone.now().strftime('%Y%m%d'))
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title = Paragraph('Spare Parts Inventory Report', styles['Title'])
    elements.append(title)
    
    # Add table
    headers = ['Code', 'Item', 'Current Inventory']
    data = [headers]
    
    for spare_part in spare_parts:
        row = [
            spare_part.code,
            spare_part.item,
            str(spare_part.inventory)
        ]
        data.append(row)
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 14),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 12),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response

@login_required
def machine_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        serial_number = request.POST.get('serial_number')
        location = request.POST.get('location')
        status = request.POST.get('status')
        
        if not all([name, serial_number, location, status]):
            messages.error(request, 'All fields are required')
            return redirect('maintenance:machine_list')
            
        try:
            machine = Machine.objects.create(
                name=name,
                serial_number=serial_number,
                location=location,
                status=status
            )
            messages.success(request, 'Machine added successfully')
            return redirect('maintenance:machine_list')
        except Exception as e:
            if 'UNIQUE constraint failed' in str(e):
                messages.error(request, f'Serial number {serial_number} is already in use. Please use a different serial number.')
            else:
                messages.error(request, 'An error occurred while adding the machine. Please try again.')
            return redirect('maintenance:machine_add')
    return render(request, 'maintenance/machine_add.html')

@login_required
def add_spare_part(request):
    if request.method == 'POST':
        form = SparePartForm(request.POST)
        if form.is_valid():
            spare_part = form.save()
            messages.success(request, 'Spare part added successfully.')
            return redirect('maintenance:spare_parts_list')
    else:
        form = SparePartForm()
    
    context = {
        'form': form,
        'title': 'Add Spare Part'
    }
    return render(request, 'maintenance/spare_part_form.html', context)

@login_required
def edit_spare_part(request, pk):
    spare_part = get_object_or_404(SparePart, pk=pk)
    
    if request.method == 'POST':
        form = SparePartForm(request.POST, instance=spare_part)
        if form.is_valid():
            form.save()
            messages.success(request, 'Spare part updated successfully.')
            return redirect('maintenance:spare_parts_list')
    else:
        form = SparePartForm(instance=spare_part)
    
    context = {
        'form': form,
        'title': 'Edit Spare Part'
    }
    return render(request, 'maintenance/spare_part_form.html', context)

@login_required
def import_spare_parts(request):
    if request.method == 'POST':
        form = SparePartImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            try:
                # Process the Excel file and create spare parts
                df = pd.read_excel(excel_file)
                for _, row in df.iterrows():
                    SparePart.objects.create(
                        code=row['Code'],
                        item=row['Item'],
                        inventory=row['Inventory']
                    )
                messages.success(request, 'Spare parts imported successfully.')
                return redirect('maintenance:spare_parts_list')
            except Exception as e:
                messages.error(request, f'Error importing spare parts: {str(e)}')
    else:
        form = SparePartImportForm()
    
    context = {
        'form': form,
        'title': 'Import Spare Parts'
    }
    return render(request, 'maintenance/import_spare_parts.html', context)

@login_required
def export_data(request, file_type):
    if file_type == 'excel':
        # Create a new workbook
        workbook = xlsxwriter.Workbook('spare_parts.xlsx')
        worksheet = workbook.add_worksheet()
        
        # Add headers
        headers = ['Code', 'Item', 'Inventory']
        row = 0
        col = 0
        
        # Write headers
        for header in headers:
            worksheet.write(row, col, header)
            col += 1
        
        # Write data
        row = 1
        spare_parts = SparePart.objects.all()
        for spare_part in spare_parts:
            worksheet.write(row, 0, spare_part.code)
            worksheet.write(row, 1, spare_part.item)
            worksheet.write(row, 2, spare_part.inventory)
            row += 1
        
        workbook.close()
        
        # Serve the file
        with open('spare_parts.xlsx', 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=spare_parts.xlsx'
            return response
    
    elif file_type == 'pdf':
        # Create the HttpResponse object with the appropriate PDF headers.
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="spare_parts.pdf"'
        
        # Create the PDF object, using the response object as its "file."
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        
        # Add title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "Spare Parts Inventory")
        
        # Add date
        p.setFont("Helvetica", 12)
        p.drawString(50, height - 70, f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Add table header
        headers = ['Code', 'Item', 'Inventory']
        y = height - 100
        for header in headers:
            p.drawString(50, y, header)
            y -= 20
        
        # Add data
        spare_parts = SparePart.objects.all()
        for spare_part in spare_parts:
            y -= 20
            if y < 50:
                p.showPage()
                y = height - 50
            
            p.drawString(50, y, spare_part.code)
            p.drawString(150, y, spare_part.item)
            p.drawString(350, y, str(spare_part.inventory))
        
        # Close the PDF object cleanly, and we're done.
        p.showPage()
        p.save()
        return response
    
    return HttpResponseBadRequest("Invalid file type")

@login_required
def spare_part_transactions(request, pk):
    spare_part = get_object_or_404(SparePart, id=pk)
    filter_form = TransactionFilterForm(request.GET)
    
    transactions = SparePartTransaction.objects.filter(spare_part=spare_part).order_by('-transaction_date')
    
    if filter_form.is_valid():
        start_date = filter_form.cleaned_data.get('start_date')
        end_date = filter_form.cleaned_data.get('end_date')
        
        if start_date:
            transactions = transactions.filter(transaction_date__gte=start_date)
        if end_date:
            transactions = transactions.filter(transaction_date__lte=end_date)
    
    export_format = filter_form.cleaned_data.get('export_format')
    if export_format:
        if export_format == 'excel':
            return export_transactions_excel(request, transactions)
        elif export_format == 'pdf':
            return export_transactions_pdf(request, transactions)
    
    context = {
        'spare_part': spare_part,
        'transactions': transactions,
        'filter_form': filter_form
    }
    return render(request, 'maintenance/spare_part_transactions.html', context)

@login_required
def delete_spare_part(request, spare_part_id):
    spare_part = get_object_or_404(SparePart, id=spare_part_id)
    
    if request.method == 'POST':
        spare_part.delete()
        messages.success(request, f'Spare part {spare_part.code} deleted successfully')
        return redirect('maintenance:spare_parts_list')
    
    return render(request, 'maintenance/delete_spare_part.html', {
        'spare_part': spare_part
    })

@login_required
def delete_transaction(request, transaction_id):
    transaction = get_object_or_404(SparePartTransaction, id=transaction_id)
    spare_part = transaction.spare_part
    
    if request.method == 'POST':
        # Update inventory based on transaction type
        if transaction.transaction_type == 'add':
            spare_part.inventory -= transaction.quantity
        else:  # remove
            spare_part.inventory += transaction.quantity
        spare_part.save()
        
        transaction.delete()
        messages.success(request, f'Transaction {transaction.id} deleted successfully')
        return redirect('maintenance:spare_part_transactions', spare_part_id=spare_part.id)
    
    return render(request, 'maintenance/delete_transaction.html', {
        'transaction': transaction,
        'spare_part': spare_part
    })

@login_required
def machine_detail(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    faults = Fault.objects.filter(machine=machine).order_by('-reported_at')
    return render(request, 'maintenance/machine_detail.html', {
        'machine': machine,
        'faults': faults
    })

@login_required
def machine_edit(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    
    if request.method == 'POST':
        machine.name = request.POST.get('name')
        machine.serial_number = request.POST.get('serial_number')
        machine.location = request.POST.get('location')
        machine.status = request.POST.get('status')
        
        if not all([machine.name, machine.serial_number, machine.location, machine.status]):
            messages.error(request, 'All fields are required')
            return redirect('maintenance:machine_detail', pk=pk)
            
        machine.save()
        messages.success(request, 'Machine updated successfully')
        return redirect('maintenance:machine_detail', pk=pk)
    
    return render(request, 'maintenance/machine_edit.html', {'machine': machine})

@login_required
def fault_list(request):
    # Get filter parameters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    machine_id = request.GET.get('machine')
    filter_type = request.GET.get('filter_type', 'all')  # 'all' or 'machine'

    # Start with all faults
    faults = Fault.objects.all()

    # Apply date filter if provided
    if start_date:
        faults = faults.filter(fault_date__gte=start_date)
    if end_date:
        faults = faults.filter(fault_date__lte=end_date)

    # Apply machine filter if provided and filter_type is 'machine'
    if machine_id and filter_type == 'machine':
        faults = faults.filter(machine_id=machine_id)
    elif filter_type == 'all':
        # If filter_type is 'all', ignore machine filter
        pass
    else:
        # If no machine is selected, show all machines
        faults = faults.all()

    # Order by fault_date descending
    faults = faults.order_by('-fault_date')

    # Get all machines for the filter dropdown
    machines = Machine.objects.all()

    return render(request, 'maintenance/fault_list.html', {
        'faults': faults,
        'machines': machines,
        'start_date': start_date,
        'end_date': end_date,
        'machine_id': machine_id,
        'filter_type': filter_type
    })

@login_required
def export_faults_excel(request):
    # Get filter parameters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    machine_id = request.GET.get('machine')
    filter_type = request.GET.get('filter_type', 'all')

    # Filter faults based on parameters
    faults = Fault.objects.all()
    if start_date:
        faults = faults.filter(fault_date__gte=start_date)
    if end_date:
        faults = faults.filter(fault_date__lte=end_date)
    if machine_id and filter_type == 'machine':
        faults = faults.filter(machine_id=machine_id)
    
    # Create Excel file
    output = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    output['Content-Disposition'] = 'attachment; filename=faults_report.xlsx'
    
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    
    # Add headers
    headers = ['Machine', 'Location', 'Fault Date', 'Reported At', 'Severity', 'Reported By', 'Status']
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header)
    
    # Add data
    row_num = 1
    for fault in faults:
        worksheet.write(row_num, 0, fault.machine.name)
        worksheet.write(row_num, 1, fault.location)
        worksheet.write(row_num, 2, fault.fault_date.strftime('%Y-%m-%d'))
        worksheet.write(row_num, 3, fault.reported_at.strftime('%Y-%m-%d %H:%M'))
        worksheet.write(row_num, 4, fault.get_severity_display())
        worksheet.write(row_num, 5, fault.reported_by.username)
        worksheet.write(row_num, 6, fault.get_status_display())
        row_num += 1
    
    workbook.close()
    return output

@login_required
def export_faults_pdf(request):
    # Get filter parameters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    machine_id = request.GET.get('machine')
    filter_type = request.GET.get('filter_type', 'all')

    # Filter faults based on parameters
    faults = Fault.objects.all()
    if start_date:
        faults = faults.filter(fault_date__gte=start_date)
    if end_date:
        faults = faults.filter(fault_date__lte=end_date)
    if machine_id and filter_type == 'machine':
        faults = faults.filter(machine_id=machine_id)
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=faults_report.pdf'
    
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    
    # Add title
    elements.append(Paragraph('Faults Report', styles['Heading1']))
    elements.append(Paragraph(f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
    
    # Add table header
    headers = ['Machine', 'Location', 'Fault Date', 'Reported At', 'Severity', 'Reported By', 'Status']
    data = [headers]
    
    # Add data rows
    for fault in faults:
        data.append([
            fault.machine.name,
            fault.location,
            fault.fault_date.strftime('%Y-%m-%d'),
            fault.reported_at.strftime('%Y-%m-%d %H:%M'),
            fault.get_severity_display(),
            fault.reported_by.username,
            fault.get_status_display()
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 14),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 12),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

@login_required
def fault_resolve(request, pk):
    fault = get_object_or_404(Fault, pk=pk)
    
    if request.method == 'POST':
        resolution_notes = request.POST.get('resolution_notes', '')
        if resolution_notes:
            fault.resolution_notes = resolution_notes
            fault.resolved_by = request.user
            fault.resolution_date = timezone.now()
            fault.status = 'resolved'
            fault.save()
            messages.success(request, 'Fault resolved successfully')
        else:
            messages.error(request, 'Please provide resolution notes')
        return redirect('maintenance:fault_list')
    
    return render(request, 'maintenance/fault_resolve.html', {'fault': fault})

@login_required
def fault_close(request, pk):
    fault = get_object_or_404(Fault, pk=pk)
    
    if request.method == 'POST':
        fault.status = 'closed'
        fault.save()
        messages.success(request, 'Fault closed successfully')
        return redirect('maintenance:fault_list')
    
    return render(request, 'maintenance/fault_close.html', {'fault': fault})

@login_required
def fault_delete(request, pk):
    fault = get_object_or_404(Fault, pk=pk)
    
    if request.method == 'POST':
        fault.delete()
        messages.success(request, 'Fault deleted successfully')
        return redirect('maintenance:fault_list')
    
    return render(request, 'maintenance/fault_delete.html', {'fault': fault})

@login_required
def fault_add(request):
    if request.method == 'POST':
        machine_id = request.POST.get('machine')
        location = request.POST.get('location')
        description = request.POST.get('description')
        severity = request.POST.get('severity')
        fault_date = request.POST.get('fault_date')
        
        if not all([machine_id, location, description, severity]):
            messages.error(request, 'All fields are required')
            return redirect('maintenance:fault_add')
            
        if fault_date:
            try:
                from datetime import datetime
                fault_date = datetime.strptime(fault_date, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'Invalid date format. Please use YYYY-MM-DD')
                return redirect('maintenance:fault_add')
        
        machine = get_object_or_404(Machine, pk=machine_id)
        fault = Fault.objects.create(
            machine=machine,
            location=location,
            description=description,
            severity=severity,
            fault_date=fault_date,
            reported_by=request.user
        )
        
        # Handle uploaded photos
        for i in range(1, 6):  # Up to 5 photos
            photo_key = f'photo_{i}'
            if request.FILES.get(photo_key):
                caption_key = f'caption_{i}'
                caption = request.POST.get(caption_key, '')
                FaultPhoto.objects.create(
                    fault=fault,
                    image=request.FILES[photo_key],
                    caption=caption
                )
        
        messages.success(request, 'Fault reported successfully')
        return redirect('maintenance:fault_list')
    
    machines = Machine.objects.all()
    return render(request, 'maintenance/fault_add.html', {'machines': machines})

@login_required
def fault_detail(request, pk):
    fault = get_object_or_404(Fault, pk=pk)
    photos = FaultPhoto.objects.filter(fault=fault).order_by('created_at')
    return render(request, 'maintenance/fault_detail.html', {
        'fault': fault,
        'photos': photos
    })

@login_required
def user_list(request):
    users = User.objects.all()
    return render(request, 'maintenance/user_list.html', {'users': users})

@login_required
def user_add(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        role = request.POST.get('role')
        department = request.POST.get('department')
        phone_number = request.POST.get('phone_number')
        
        if not all([username, password, email, first_name, last_name, role]):
            messages.error(request, 'All fields are required')
            return redirect('maintenance:user_list')
            
        try:
            user = User.objects.create_user(
                username=username,
                password=password,
                email=email,
                first_name=first_name,
                last_name=last_name
            )
            
            user_profile = UserProfile.objects.create(
                user=user,
                role=UserRole.objects.get(name=role),
                department=department,
                phone_number=phone_number
            )
            
            messages.success(request, 'User added successfully')
            return redirect('maintenance:user_list')
        except Exception as e:
            messages.error(request, f'Error creating user: {str(e)}')
            return redirect('maintenance:user_list')
    
    roles = UserRole.objects.all()
    return render(request, 'maintenance/user_add.html', {'roles': roles})

@login_required
def user_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    user_profile = get_object_or_404(UserProfile, user=user)
    
    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user_profile.role = UserRole.objects.get(name=request.POST.get('role'))
        user_profile.department = request.POST.get('department')
        user_profile.phone_number = request.POST.get('phone_number')
        
        if not all([user.username, user.email, user.first_name, user.last_name, user_profile.role]):
            messages.error(request, 'All fields are required')
            return redirect('maintenance:user_edit', pk=pk)
            
        user.save()
        user_profile.save()
        messages.success(request, 'User updated successfully')
        return redirect('maintenance:user_list')
    
    roles = UserRole.objects.all()
    return render(request, 'maintenance/user_edit.html', {
        'user': user,
        'user_profile': user_profile,
        'roles': roles
    })

@login_required
def user_profile(request):
    user = request.user
    user_profile = get_object_or_404(UserProfile, user=user)
    
    if request.method == 'POST':
        user.email = request.POST.get('email')
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user_profile.department = request.POST.get('department')
        user_profile.phone_number = request.POST.get('phone_number')
        
        if not all([user.email, user.first_name, user.last_name]):
            messages.error(request, 'Email, first name, and last name are required')
            return redirect('maintenance:user_profile')
            
        user.save()
        user_profile.save()
        messages.success(request, 'Profile updated successfully')
        return redirect('maintenance:user_profile')
    
    return render(request, 'maintenance/user_profile.html', {
        'user': user,
        'user_profile': user_profile
    })

@login_required
def user_detail(request, pk):
    user = get_object_or_404(User, pk=pk)
    user_profile = get_object_or_404(UserProfile, user=user)
    
    context = {
        'user': user,
        'user_profile': user_profile,
        'title': f'User Profile - {user.username}'
    }
    return render(request, 'maintenance/user_detail.html', context)

@login_required
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'User deleted successfully.')
        return redirect('maintenance:users')
    return render(request, 'maintenance/user_confirm_delete.html', {'user': user})
